import requests
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

url = (
    "https://prd-usta-kube-tournamentdesk-public-api.clubspark.pro/"
)


headers = {
    "User-Agent": "Mozilla/5.0",
    "Content-Type": "application/json",
}

payload = {
    "query": """
        query TournamentPublicEventData($eventId: ID!, $tournamentId: ID!) {
          tournamentPublicEventData(eventId: $eventId, tournamentId: $tournamentId)
        }
    """,
    "variables": {
        "eventId": "E2884F9E-6C3A-4DE4-B751-E27628267559",
        "tournamentId": "45779550-A952-495F-B0CF-68EA5F2DCBC0",
    },
}

response = requests.post(url, headers=headers, json=payload)

data = response.json()
roundMatchUps = data["data"]["tournamentPublicEventData"]["eventData"]["drawsData"][0][
    "structures"
][1]["roundMatchUps"]

matches_info = []


for matchUp in roundMatchUps:
    currentRound = roundMatchUps[matchUp]
    for match in currentRound:
        if match["matchUpStatus"] == "BYE":
            matches_info.append(
                {
                    "Round": match["roundName"],
                    "Player1": "",
                    "Player2": "",
                    "Winner": "BYE",
                    "Player1URL": "",
                    "Player2URL": "",
                }
            )
        elif match["matchUpStatus"] == "WALKOVER":
            matches_info.append(
                {
                    "Round": match["roundName"],
                    "Player1": "",
                    "Player2": "",
                    "Winner": "INJ",
                    "Player1URL": "",
                    "Player2URL": "",
                }
            )
        elif match["matchUpStatus"] != "COMPLETED" and match["matchUpStatus"] != "RETIRED":
            matches_info.append(
                {
                    "Round": match["roundName"],
                    "Player1": "",
                    "Player2": "",
                    "Winner": "MISC",
                    "Player1URL": "",
                    "Player2URL": "",
                }
            )
        else:
            player1 = (
                "https://www.usta.com/en/home/play/player-search/profile.html#?uaid="
                + str(
                    match["sides"][0]["participant"]["person"]["personOtherIds"][0][
                        "personId"
                    ]
                )
            )
            player2 = (
                "https://www.usta.com/en/home/play/player-search/profile.html#?uaid="
                + str(
                    match["sides"][1]["participant"]["person"]["personOtherIds"][0][
                        "personId"
                    ]
                )
            )

            matches_info.append(
                {
                    "Round": match["roundName"],
                    "Player1": match["sides"][0]["participant"]["person"][
                        "standardFamilyName"
                    ]
                    + ", "
                    + match["sides"][0]["participant"]["person"]["standardGivenName"],
                    "Player2": match["sides"][1]["participant"]["person"][
                        "standardFamilyName"
                    ]
                    + ", "
                    + match["sides"][1]["participant"]["person"]["standardGivenName"],
                    "Winner": match["winningSide"],
                    "Player1URL": player1,
                    "Player2URL": player2,
                }
            )
df = pd.DataFrame(matches_info)

file_name = "output_B16_Consolation.xlsx"
df[["Round", "Player1", "Player2", "Winner"]].to_excel(
    file_name, index=False, engine="openpyxl"
)

workbook = load_workbook(file_name)
sheet = workbook.active


for row_num in range(2, sheet.max_row + 1):  #starting from 2 to skip header
    player1_cell = sheet.cell(
        row=row_num, column=2
    ) 
    player1_url = df.loc[
        row_num - 2, "Player1URL"
    ]
    player1_cell.hyperlink = player1_url
    player1_cell.font = Font(color="0000FF", underline="single")

    player2_cell = sheet.cell(
        row=row_num, column=3
    ) 
    player2_url = df.loc[row_num - 2, "Player2URL"]
    player2_cell.hyperlink = player2_url
    player2_cell.font = Font(color="0000FF", underline="single")

#Excel
workbook.save(file_name)
